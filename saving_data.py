import argparse
import json
import os
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook


CONFIG_FILENAME = "config.json"
SHEET_NAME = "sales_leads"
FIELDNAMES: List[str] = [
    "visitor_name",
    "title",
    "company",
    "interests_of_solutions",
    "interested_in_pilot",
    "email",
    "phone_number",
    "next_steps",
]


def _load_target_excel_path() -> Path:
    """
    Load the target Excel file path from config.json in the current directory.

    Expects a JSON object with a "target_excel_file" key.
    Raises an informative error if the configuration or file is invalid.
    """
    config_path = Path(CONFIG_FILENAME)
    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file '{CONFIG_FILENAME}' not found.")

    try:
        config_data = json.loads(config_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON in '{CONFIG_FILENAME}'.") from exc

    target_value = config_data.get("target_excel_file")
    if not target_value:
        raise KeyError(
            "Missing 'target_excel_file' key in configuration file "
            f"'{CONFIG_FILENAME}'."
        )

    target_path = Path(target_value)
    if not target_path.exists():
        raise FileNotFoundError(
            f"Configured Excel file does not exist: '{target_path}'. "
            "Please create the file before running this script."
        )

    if not os.access(target_path, os.W_OK):
        raise PermissionError(
            f"No write access to configured Excel file: '{target_path}'."
        )

    return target_path


def _get_or_create_sheet(path: Path):
    workbook = load_workbook(path)
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
    else:
        sheet = workbook.create_sheet(SHEET_NAME)
    return workbook, sheet


def save_lead(**lead: Any) -> None:
    """
    Append a single sales lead row to the Excel file in the 'sales_leads' sheet.

    The workbook and sheet are created with a header row if they do not exist.
    """
    path = _load_target_excel_path()
    workbook, sheet = _get_or_create_sheet(path)

    # Add header row if the sheet is empty
    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        sheet.append(FIELDNAMES)

    row: Dict[str, Any] = {name: lead.get(name, "") for name in FIELDNAMES}
    sheet.append([row[name] for name in FIELDNAMES])

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Save sales lead information to a local Excel file."
    )
    parser.add_argument("--visitor_name", required=True, help="Name of the visitor.")
    parser.add_argument("--title", required=True, help="Job title of the visitor.")
    parser.add_argument("--company", required=True, help="Company name.")
    parser.add_argument(
        "--interests_of_solutions",
        required=True,
        help="Solutions or products the visitor is interested in.",
    )
    parser.add_argument(
        "--interested_in_pilot",
        required=True,
        help="Whether the visitor is interested in a pilot (e.g., yes/no).",
    )
    parser.add_argument("--email", required=True, help="Email address.")
    parser.add_argument("--phone_number", required=True, help="Phone number.")
    parser.add_argument(
        "--next_steps",
        required=True,
        help="Agreed next steps or follow-up actions.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    save_lead(
        visitor_name=args.visitor_name,
        title=args.title,
        company=args.company,
        interests_of_solutions=args.interests_of_solutions,
        interested_in_pilot=args.interested_in_pilot,
        email=args.email,
        phone_number=args.phone_number,
        next_steps=args.next_steps,
    )


if __name__ == "__main__":
    main()
