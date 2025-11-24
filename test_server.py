import unittest
import os
import sys
import json
import time
import subprocess
import signal
import asyncio
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


class TestData_storageServer(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.server_process = None
        cls.port = int(os.getenv('MCP_DATA_STORAGE_PORT', 7905))
        cls.server_path = "server.py"

    def setUp(self):
        # Start server before each test
        self.server_process = subprocess.Popen(
            [sys.executable, self.server_path, "start", "--json"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        time.sleep(3)  # Wait for server to start

    def tearDown(self):
        # Stop server after each test
        if self.server_process:
            try:
                # Try to stop gracefully first
                subprocess.run([sys.executable, self.server_path, "stop", "--json"], 
                             timeout=5, capture_output=True)
            except:
                pass
            finally:
                self.server_process.stdout.close()
                self.server_process.stderr.close()
                self.server_process.terminate()
                self.server_process.wait(timeout=5)

    def test_server_version(self):
        """Test that server version command works"""
        result = subprocess.run(
            [sys.executable, self.server_path, "version", "--json"],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(result.returncode, 0)
        
        # Parse JSON response - handle mixed output robustly
        response = self._parse_json_from_output(result.stdout)
        self.assertTrue(response.get('success', False))
        self.assertIn('version', response.get('data', {}))

    def test_server_status(self):
        """Test that server status command works"""
        result = subprocess.run(
            [sys.executable, self.server_path, "status", "--json"],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(result.returncode, 0)
        
        # Parse JSON response - handle mixed output robustly
        response = self._parse_json_from_output(result.stdout)
        self.assertIn('success', response)
        self.assertIn('message', response)

    def test_server_ping(self):
        """Test that server ping command works"""
        result = subprocess.run(
            [sys.executable, self.server_path, "ping", "--json"],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(result.returncode, 0)
        
        # Parse JSON response - handle mixed output robustly
        response = self._parse_json_from_output(result.stdout)
        self.assertIn('success', response)
        self.assertIn('message', response)

    def _parse_json_from_output(self, output):
        """
        Robustly parse JSON from mixed output that may contain logging messages.
        Looks for lines that start with '{' or '[' and attempts to parse them as JSON.
        """
        lines = output.strip().split('\n')
        
        # Try to find and parse JSON lines
        for line in lines:
            line_stripped = line.strip()
            # Only attempt JSON parsing if line looks like JSON (starts with '{' or '[')
            if line_stripped.startswith('{') or line_stripped.startswith('['):
                try:
                    return json.loads(line_stripped)
                except json.JSONDecodeError:
                    continue
        
        # If no valid JSON found, try parsing the entire output as a fallback
        try:
            return json.loads(output.strip())
        except json.JSONDecodeError as e:
            self.fail(f"Could not parse JSON from output: {output}\nError: {e}")


class TestSaveSalesLead(unittest.TestCase):
    def setUp(self):
        # Create a temporary directory with its own config and Excel file
        self._orig_cwd = os.getcwd()
        self._tmpdir = tempfile.TemporaryDirectory()
        os.chdir(self._tmpdir.name)

        self.excel_path = Path(self._tmpdir.name) / "sales_leads_test.xlsx"

        # Create an empty workbook so the file exists and is writable
        wb = Workbook()
        wb.save(self.excel_path)

        # Write a config.json that points to the temporary Excel file
        config_path = Path("config.json")
        config_data = {"target_excel_file": str(self.excel_path)}
        config_path.write_text(json.dumps(config_data), encoding="utf-8")

    def tearDown(self):
        # Restore original working directory and clean up
        os.chdir(self._orig_cwd)
        self._tmpdir.cleanup()

    def test_save_sales_lead_appends_row_and_header(self):
        from server import save_sales_lead, SHEET_NAME, FIELDNAMES

        # Call the async tool
        asyncio.run(
            save_sales_lead(
                visitor_name="Jane Smith",
                title="CTO",
                company="Example Inc.",
                interests_of_solutions="Cloud storage, disaster recovery",
                interested_in_pilot="yes",
                email="jane.smith@example.com",
                phone_number="+1-555-0200",
                next_steps="Send proposal next week",
            )
        )

        # Verify the workbook contents
        workbook = load_workbook(self.excel_path)
        sheet = workbook[SHEET_NAME]

        # Header row should match FIELDNAMES
        header_values = [cell.value for cell in sheet[1]]
        assert header_values == FIELDNAMES

        # The second row should contain the saved lead data
        data_row = [cell.value for cell in sheet[2]]
        expected = [
            "Jane Smith",
            "CTO",
            "Example Inc.",
            "Cloud storage, disaster recovery",
            "yes",
            "jane.smith@example.com",
            "+1-555-0200",
            "Send proposal next week",
        ]
        assert data_row == expected

if __name__ == '__main__':
    unittest.main()
